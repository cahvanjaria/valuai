"""
ValuAI - Excel Reader Module
=============================
Reads a filled-in template.xlsx and returns structured data for downstream
valuation engines (DCF, CCM, NAV, sensitivity, report generation).

Author: RV Harshal Vibhakar Anjaria | IBBI/RV/03/2026/16120
Project: ValuAI - ICAI AI Hackathon 2026

CONVENTION NOTE:
  - All percentages stored as plain numbers (e.g., 7.21 means 7.21%, not 0.0721)
  - Currency unit is configurable (Lakhs/Millions/Crores per Cover!B26)
  - Computed cells (formulas) are read with their evaluated values
"""

from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime, date


# ----------------------------------------------------------------------------
# Helper functions
# ----------------------------------------------------------------------------

def _cell(ws, ref):
    """Return the value of a single cell, or None if empty."""
    value = ws[ref].value
    if value == "" or value is None:
        return None
    return value


def _row_values(ws, row_num, start_col, end_col):
    """Return a list of values from one row, columns start_col to end_col inclusive.
    Empty cells become None. Used to read year-by-year data."""
    values = []
    for col in range(start_col, end_col + 1):
        v = ws.cell(row=row_num, column=col).value
        if v == "":
            v = None
        values.append(v)
    return values


def _to_iso_date(value):
    """Convert various date formats into an ISO-format string (YYYY-MM-DD).
    Handles datetime objects, date objects, and strings. Returns None if blank."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


# ----------------------------------------------------------------------------
# Metadata reader (Cover sheet)
# ----------------------------------------------------------------------------

def read_metadata(wb):
    """Extract engagement and methodology metadata from the Cover sheet."""
    ws = wb["Cover"]

    return {
        "valuer_name": _cell(ws, "B4"),
        "valuer_ibbi_reg_no": _cell(ws, "B5"),
        "valuer_asset_class": _cell(ws, "B6"),
        "valuer_address": _cell(ws, "B7"),
        "client_name": _cell(ws, "B10"),
        "client_cin": _cell(ws, "B11"),
        "client_address": _cell(ws, "B12"),
        "date_of_appointment": _to_iso_date(_cell(ws, "B13")),
        "valuation_date": _to_iso_date(_cell(ws, "B14")),
        "date_of_report": _to_iso_date(_cell(ws, "B15")),
        "purpose": _cell(ws, "B16"),
        "statutory_reference": _cell(ws, "B17"),
        "base_of_value": _cell(ws, "B20"),
        "premise_of_value": _cell(ws, "B21"),
        "valuation_standard_ref": _cell(ws, "B22"),
        "currency": _cell(ws, "B25"),
        "units": _cell(ws, "B26"),
        "historical_years": _cell(ws, "B29"),
        "forecast_years": _cell(ws, "B30"),
        "stub_period": _cell(ws, "B31"),
        "stub_months": _cell(ws, "B32"),
        "discounting_convention": _cell(ws, "B35"),
        "wacc_input_mode": _cell(ws, "B36"),
        "terminal_value_method": _cell(ws, "B37"),
        "use_case_type": _cell(ws, "B40"),
        "investment_amount": _cell(ws, "B41"),
        "pre_investment_shares": _cell(ws, "B42"),
        "new_shares_issued": _cell(ws, "B43"),
        "post_investment_shares": _cell(ws, "B44"),
        "discount_rate_method": _cell(ws, "B45"),
        "reconciliation_tolerance": _cell(ws, "B46"),
    }


# ----------------------------------------------------------------------------
# Historical P&L reader
# ----------------------------------------------------------------------------

def read_historical_pl(wb, num_historical_years):
    """Read historical Profit & Loss data (columns B..F = FY-4 to FY-0 Latest)."""
    ws = wb["Historical_PL"]
    START_COL, END_COL = 2, 6

    return {
        "year_endings": [_to_iso_date(v) for v in _row_values(ws, 4, START_COL, END_COL)],
        "revenue_from_operations": _row_values(ws, 7, START_COL, END_COL),
        "other_income": _row_values(ws, 8, START_COL, END_COL),
        "total_revenue": _row_values(ws, 9, START_COL, END_COL),
        "cost_of_goods_sold": _row_values(ws, 12, START_COL, END_COL),
        "employee_benefits": _row_values(ws, 13, START_COL, END_COL),
        "other_operating_expenses": _row_values(ws, 14, START_COL, END_COL),
        "total_operating_costs": _row_values(ws, 15, START_COL, END_COL),
        "ebitda": _row_values(ws, 17, START_COL, END_COL),
        "depreciation": _row_values(ws, 18, START_COL, END_COL),
        "ebit": _row_values(ws, 19, START_COL, END_COL),
        "finance_cost": _row_values(ws, 20, START_COL, END_COL),
        "pbt": _row_values(ws, 21, START_COL, END_COL),
        "tax_expense": _row_values(ws, 22, START_COL, END_COL),
        "pat": _row_values(ws, 23, START_COL, END_COL),
        "num_years_configured": num_historical_years,
    }


# ----------------------------------------------------------------------------
# Historical Balance Sheet reader
# ----------------------------------------------------------------------------

def read_historical_bs(wb, num_historical_years):
    """Read historical Balance Sheet data."""
    ws = wb["Historical_BS"]
    START_COL, END_COL = 2, 6

    return {
        "as_on_dates": [_to_iso_date(v) for v in _row_values(ws, 4, START_COL, END_COL)],
        "equity_share_capital": _row_values(ws, 8, START_COL, END_COL),
        "reserves_and_surplus": _row_values(ws, 9, START_COL, END_COL),
        "total_shareholders_funds": _row_values(ws, 10, START_COL, END_COL),
        "preference_share_capital": _row_values(ws, 11, START_COL, END_COL),
        "long_term_borrowings": _row_values(ws, 14, START_COL, END_COL),
        "deferred_tax_liabilities": _row_values(ws, 15, START_COL, END_COL),
        "other_non_current_liabilities": _row_values(ws, 16, START_COL, END_COL),
        "total_non_current_liabilities": _row_values(ws, 17, START_COL, END_COL),
        "short_term_borrowings": _row_values(ws, 20, START_COL, END_COL),
        "trade_payables": _row_values(ws, 21, START_COL, END_COL),
        "other_current_liabilities": _row_values(ws, 22, START_COL, END_COL),
        "short_term_provisions": _row_values(ws, 23, START_COL, END_COL),
        "total_current_liabilities": _row_values(ws, 24, START_COL, END_COL),
        "total_equity_and_liabilities": _row_values(ws, 26, START_COL, END_COL),
        "property_plant_equipment": _row_values(ws, 31, START_COL, END_COL),
        "capital_work_in_progress": _row_values(ws, 32, START_COL, END_COL),
        "intangible_assets": _row_values(ws, 33, START_COL, END_COL),
        "long_term_loans_advances": _row_values(ws, 34, START_COL, END_COL),
        "other_non_current_assets": _row_values(ws, 35, START_COL, END_COL),
        "non_current_investments": _row_values(ws, 36, START_COL, END_COL),
        "deferred_tax_assets": _row_values(ws, 37, START_COL, END_COL),
        "total_non_current_assets": _row_values(ws, 38, START_COL, END_COL),
        "current_investments": _row_values(ws, 41, START_COL, END_COL),
        "inventories": _row_values(ws, 42, START_COL, END_COL),
        "trade_receivables": _row_values(ws, 43, START_COL, END_COL),
        "cash_and_equivalents": _row_values(ws, 44, START_COL, END_COL),
        "bank_balances_other": _row_values(ws, 45, START_COL, END_COL),
        "short_term_loans_advances": _row_values(ws, 46, START_COL, END_COL),
        "other_current_assets": _row_values(ws, 47, START_COL, END_COL),
        "total_current_assets": _row_values(ws, 48, START_COL, END_COL),
        "total_assets": _row_values(ws, 50, START_COL, END_COL),
        "tally_check": _row_values(ws, 52, START_COL, END_COL),
        "num_years_configured": num_historical_years,
    }


# ----------------------------------------------------------------------------
# Projections reader (the most important sheet)
# ----------------------------------------------------------------------------

def read_projections(wb, num_forecast_years, has_stub):
    """Read projection (forecast) data. Columns B..L: B=Stub, C..L=FY+1..FY+10."""
    ws = wb["Projections"]
    START_COL, END_COL = 2, 12

    return {
        "year_endings": [_to_iso_date(v) for v in _row_values(ws, 4, START_COL, END_COL)],
        "revenue_from_operations": _row_values(ws, 7, START_COL, END_COL),
        "other_income": _row_values(ws, 8, START_COL, END_COL),
        "total_revenue": _row_values(ws, 9, START_COL, END_COL),
        "cost_of_goods_sold": _row_values(ws, 11, START_COL, END_COL),
        "employee_benefits": _row_values(ws, 12, START_COL, END_COL),
        "other_operating_expenses": _row_values(ws, 13, START_COL, END_COL),
        "total_operating_costs": _row_values(ws, 14, START_COL, END_COL),
        "ebitda": _row_values(ws, 16, START_COL, END_COL),
        "depreciation": _row_values(ws, 17, START_COL, END_COL),
        "ebit": _row_values(ws, 18, START_COL, END_COL),
        "effective_tax_rate_pct": _row_values(ws, 21, START_COL, END_COL),
        "tax_on_ebit": _row_values(ws, 22, START_COL, END_COL),
        "nopat": _row_values(ws, 23, START_COL, END_COL),
        "change_in_working_capital": _row_values(ws, 26, START_COL, END_COL),
        "capex": _row_values(ws, 27, START_COL, END_COL),
        "opening_debt": _row_values(ws, 30, START_COL, END_COL),
        "new_debt_drawn": _row_values(ws, 31, START_COL, END_COL),
        "principal_repayment": _row_values(ws, 32, START_COL, END_COL),
        "closing_debt": _row_values(ws, 33, START_COL, END_COL),
        "interest_rate_pct": _row_values(ws, 35, START_COL, END_COL),
        "interest_expense": _row_values(ws, 36, START_COL, END_COL),
        "fcff_nopat": _row_values(ws, 40, START_COL, END_COL),
        "fcff_depreciation": _row_values(ws, 41, START_COL, END_COL),
        "fcff_change_in_wc": _row_values(ws, 42, START_COL, END_COL),
        "fcff_capex": _row_values(ws, 43, START_COL, END_COL),
        "fcff": _row_values(ws, 44, START_COL, END_COL),
        "fcfe_fcff": _row_values(ws, 47, START_COL, END_COL),
        "fcfe_interest_tax_shield": _row_values(ws, 48, START_COL, END_COL),
        "fcfe_principal_repayment": _row_values(ws, 49, START_COL, END_COL),
        "fcfe_new_debt_drawn": _row_values(ws, 50, START_COL, END_COL),
        "fcfe_other_adjustments": _row_values(ws, 51, START_COL, END_COL),
        "fcfe": _row_values(ws, 52, START_COL, END_COL),
        "cash_and_equivalents_valuation_date": _cell(ws, "B55"),
        "current_investments_valuation_date": _cell(ws, "B56"),
        "total_debt_valuation_date": _cell(ws, "B57"),
        "preference_share_capital_valuation_date": _cell(ws, "B58"),
        "other_equity_bridge_adjustments": _cell(ws, "B59"),
        "shares_outstanding_millions": _cell(ws, "B62"),
        "face_value_per_share": _cell(ws, "B63"),
        "num_forecast_years_configured": num_forecast_years,
        "has_stub_period": has_stub,
    }


# ----------------------------------------------------------------------------
# WACC inputs reader
# ----------------------------------------------------------------------------

def read_wacc_inputs(wb):
    """Read WACC build-up: CAPM components, cost of debt, weights, computed WACC."""
    ws = wb["WACC_Inputs"]

    return {
        "risk_free_rate_pct": _cell(ws, "B6"),
        "equity_risk_premium_pct": _cell(ws, "B7"),
        "beta": _cell(ws, "B8"),
        "beta_source": _cell(ws, "B9"),
        "base_capm_ke_pct": _cell(ws, "B10"),
        "csrp_pct": _cell(ws, "B12"),
        "adjusted_ke_pct": _cell(ws, "B15"),
        "direct_ke_pct": _cell(ws, "B18"),
        "pre_tax_kd_pct": _cell(ws, "B22"),
        "effective_tax_rate_pct": _cell(ws, "B23"),
        "post_tax_kd_pct": _cell(ws, "B24"),
        "weight_equity_pct": _cell(ws, "B27"),
        "weight_debt_pct": _cell(ws, "B28"),
        "weights_total_pct": _cell(ws, "B29"),
        "effective_ke_pct": _cell(ws, "B32"),
        "effective_kd_pct": _cell(ws, "B33"),
        "wacc_pct": _cell(ws, "B34"),
        "rf_source": _cell(ws, "B37"),
        "erp_source": _cell(ws, "B38"),
        "beta_source_note": _cell(ws, "B39"),
        "csrp_justification": _cell(ws, "B40"),
        "other_notes": _cell(ws, "B41"),
    }


# ----------------------------------------------------------------------------
# Other inputs reader
# ----------------------------------------------------------------------------

def read_other_inputs(wb):
    """Read terminal value, sensitivity, DLOM, DLOC, weightages, pre/post-money."""
    ws = wb["Other_Inputs"]

    return {
        "terminal_value_method": _cell(ws, "B4"),
        "perpetual_growth_rate_pct": _cell(ws, "B7"),
        "exit_ebitda_multiple": _cell(ws, "B11"),
        "sensitivity_wacc": (_cell(ws, "B15"), _cell(ws, "C15")),
        "sensitivity_terminal_growth": (_cell(ws, "B16"), _cell(ws, "C16")),
        "sensitivity_ebit_margin": (_cell(ws, "B17"), _cell(ws, "C17")),
        "sensitivity_capex": (_cell(ws, "B18"), _cell(ws, "C18")),
        "sensitivity_working_capital": (_cell(ws, "B19"), _cell(ws, "C19")),
        "dlom_apply": _cell(ws, "B24"),
        "dlom_pct": _cell(ws, "B25"),
        "dlom_justification": _cell(ws, "B27"),
        "dloc_apply": _cell(ws, "B30"),
        "dloc_pct": _cell(ws, "B31"),
        "dloc_justification": _cell(ws, "B33"),
        "weight_dcf_pct": _cell(ws, "B36"),
        "weight_ccm_pct": _cell(ws, "B37"),
        "weight_nav_pct": _cell(ws, "B38"),
        "weights_total_pct": _cell(ws, "B39"),
        "weights_reasoning": _cell(ws, "B40"),
        "use_case": _cell(ws, "B43"),
        "investment_amount": _cell(ws, "B44"),
        "pre_investment_shares": _cell(ws, "B45"),
        "new_shares_issued": _cell(ws, "B46"),
        "post_investment_shares": _cell(ws, "B47"),
    }


# ----------------------------------------------------------------------------
# Comparables reader
# ----------------------------------------------------------------------------

def read_comparables(wb):
    """Read CCM inputs: subject company metrics and peer table with multiples."""
    ws = wb["Comparables"]

    subject = {
        "revenue": _cell(ws, "B6"),
        "ebitda": _cell(ws, "B7"),
        "ebit": _cell(ws, "B8"),
        "pat": _cell(ws, "B9"),
        "book_value_equity": _cell(ws, "B10"),
    }

    peers = []
    for row in range(14, 22):
        peer = {
            "company_name": _cell(ws, f"B{row}"),
            "exchange_symbol": _cell(ws, f"C{row}"),
            "industry": _cell(ws, f"D{row}"),
            "market_cap_cr": _cell(ws, f"E{row}"),
            "enterprise_value_cr": _cell(ws, f"F{row}"),
            "ev_revenue": _cell(ws, f"G{row}"),
            "ev_ebitda": _cell(ws, f"H{row}"),
            "ev_ebit": _cell(ws, f"I{row}"),
            "pe": _cell(ws, f"J{row}"),
            "pb": _cell(ws, f"K{row}"),
            "selection_justification": _cell(ws, f"L{row}"),
        }
        if peer["company_name"] is not None:
            peers.append(peer)

    medians = {
        "ev_revenue": _cell(ws, "G22"),
        "ev_ebitda": _cell(ws, "H22"),
        "ev_ebit": _cell(ws, "I22"),
        "pe": _cell(ws, "J22"),
        "pb": _cell(ws, "K22"),
    }

    return {
        "subject_company": subject,
        "peers": peers,
        "medians": medians,
        "primary_multiple": _cell(ws, "B25"),
        "multiple_statistic": _cell(ws, "B26"),
        "custom_multiple_value": _cell(ws, "B27"),
        "ccm_dlom_pct": _cell(ws, "B29"),
        "ccm_dlom_justification": _cell(ws, "B31"),
    }


# ----------------------------------------------------------------------------
# Master orchestrator
# ----------------------------------------------------------------------------

def read_template(file_path):
    """Read a filled-in ValuAI template (.xlsx) and return all data as a
    single structured dictionary.
    """
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")

    wb = load_workbook(file_path, data_only=True)

    expected_sheets = {
        "Cover", "Historical_PL", "Historical_BS", "Projections",
        "WACC_Inputs", "Other_Inputs", "Comparables",
    }
    missing = expected_sheets - set(wb.sheetnames)
    if missing:
        raise ValueError(f"Missing required sheets: {missing}")

    metadata = read_metadata(wb)
    n_hist = metadata.get("historical_years") or 3
    n_fcst = metadata.get("forecast_years") or 5
    has_stub = (metadata.get("stub_period") or "No") == "Yes"

    return {
        "metadata": metadata,
        "historical_pl": read_historical_pl(wb, n_hist),
        "historical_bs": read_historical_bs(wb, n_hist),
        "projections": read_projections(wb, n_fcst, has_stub),
        "wacc_inputs": read_wacc_inputs(wb),
        "other_inputs": read_other_inputs(wb),
        "comparables": read_comparables(wb),
    }


# ----------------------------------------------------------------------------
# Test / demo block (runs when this file is executed directly)
# ----------------------------------------------------------------------------

if __name__ == "__main__":
    test_file = Path("assets/case_farmgas_sejal.xlsx")
    if not test_file.exists():
        print(f"Test file not found: {test_file}")
        print("Run this script from the ValuAI root directory.")
        exit(1)

    data = read_template(test_file)

    print("=" * 70)
    print("VALUAI - EXCEL READER TEST")
    print(f"Source: {test_file}")
    print("=" * 70)

    md = data["metadata"]
    print(f"\n[METADATA]")
    print(f"  Client:           {md['client_name']}")
    print(f"  Valuation Date:   {md['valuation_date']}")
    print(f"  Valuer:           {md['valuer_name']} ({md['valuer_ibbi_reg_no']})")
    stub_str = f", with {md['stub_months']}m stub" if md['stub_period'] == 'Yes' else ''
    print(f"  Period Config:    {md['historical_years']} historical + {md['forecast_years']} forecast years{stub_str}")
    print(f"  Methodology:      {md['discounting_convention']}, {md['wacc_input_mode']}, {md['terminal_value_method']}")
    print(f"  Currency/Units:   {md['currency']} {md['units']}")

    wacc = data["wacc_inputs"]
    print(f"\n[WACC BUILD-UP]")
    print(f"  Risk-Free Rate:   {wacc['risk_free_rate_pct']}%")
    print(f"  ERP:              {wacc['equity_risk_premium_pct']}%")
    print(f"  Beta:             {wacc['beta']}")
    print(f"  Base CAPM Ke:     {wacc['base_capm_ke_pct']:.4f}%")
    print(f"  CSRP:             {wacc['csrp_pct']}%")
    print(f"  Adjusted Ke:      {wacc['adjusted_ke_pct']:.4f}%")
    print(f"  Pre-tax Kd:       {wacc['pre_tax_kd_pct']}%")
    print(f"  Post-tax Kd:      {wacc['post_tax_kd_pct']:.4f}%")
    print(f"  Weights (E/D):    {wacc['weight_equity_pct']}% / {wacc['weight_debt_pct']}%")
    print(f"  WACC:             {wacc['wacc_pct']:.4f}%")

    proj = data["projections"]
    print(f"\n[PROJECTIONS - KEY METRICS]")
    print(f"  Year endings:     {proj['year_endings']}")
    print(f"  EBIT:             {[round(v, 2) if v else None for v in proj['ebit']]}")
    print(f"  FCFF:             {[round(v, 2) if v else None for v in proj['fcff']]}")
    print(f"  FCFE:             {[round(v, 2) if v else None for v in proj['fcfe']]}")
    print(f"  Shares Outstanding: {proj['shares_outstanding_millions']} millions")
    print(f"  Cash @ Val Date:  {proj['cash_and_equivalents_valuation_date']}")
    print(f"  Debt @ Val Date:  {proj['total_debt_valuation_date']}")

    other = data["other_inputs"]
    print(f"\n[OTHER INPUTS]")
    print(f"  Terminal Growth:  {other['perpetual_growth_rate_pct']}%")
    print(f"  Method Weights:   DCF {other['weight_dcf_pct']}% / CCM {other['weight_ccm_pct']}% / NAV {other['weight_nav_pct']}%")
    print(f"  DLOM Apply:       {other['dlom_apply']} ({other['dlom_pct']}%)")
    print(f"  DLOC Apply:       {other['dloc_apply']} ({other['dloc_pct']}%)")

    comp = data["comparables"]
    print(f"\n[COMPARABLES]")
    print(f"  Number of peers:  {len(comp['peers'])}")
    if comp["peers"]:
        print(f"  First peer:       {comp['peers'][0]['company_name']}")
    else:
        print(f"  (No comparables - acceptable when CCM is not used)")

    print("\n" + "=" * 70)
    print("Excel reader test complete - all 7 sheets parsed successfully.")
    print("=" * 70)